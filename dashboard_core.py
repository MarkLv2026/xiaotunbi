# -*- coding: utf-8 -*-
from __future__ import annotations
import datetime, io, re, zipfile, csv
import xml.etree.ElementTree as ET
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, Iterable, List

NS='{http://schemas.openxmlformats.org/spreadsheetml/2006/main}'
METRICS=['商品访客数','商品浏览量','商品加购人数','商品加购件数','支付买家数','支付件数','支付金额','成功退款金额']


def _load_shared(z: zipfile.ZipFile):
    if 'xl/sharedStrings.xml' not in z.namelist():
        return []
    root=ET.fromstring(z.read('xl/sharedStrings.xml'))
    return [''.join((t.text or '') for t in si.iter(NS+'t')) for si in root.findall(NS+'si')]


def _sheet_paths(z: zipfile.ZipFile):
    names=z.namelist(); out={}
    try:
        wb=ET.fromstring(z.read('xl/workbook.xml'))
        rels=ET.fromstring(z.read('xl/_rels/workbook.xml.rels'))
        rid_to_target={r.attrib['Id']:r.attrib['Target'] for r in rels}
        for s in wb.find(NS+'sheets'):
            name=s.attrib.get('name','')
            rid=s.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
            target=rid_to_target.get(rid,'')
            path='xl/'+target.lstrip('/'); path=path.replace('xl/xl/','xl/')
            if name in ('天猫数据源','京东抖音数据源') and path in names:
                out[name]=path
    except Exception:
        pass
    if not out:
        if 'xl/worksheets/sheet1.xml' in names: out['天猫数据源']='xl/worksheets/sheet1.xml'
        if 'xl/worksheets/sheet2.xml' in names: out['京东抖音数据源']='xl/worksheets/sheet2.xml'
    return out


def _col_idx(ref: str) -> int:
    m=re.match(r'([A-Z]+)', ref or '')
    if not m: return 0
    n=0
    for ch in m.group(1): n=n*26+ord(ch)-64
    return n-1


def _cell_value(c, ss):
    v=c.find(NS+'v')
    if v is None: return None
    txt=v.text
    if c.attrib.get('t')=='s':
        try: return ss[int(txt)]
        except Exception: return txt
    return txt


def _iter_rows(z: zipfile.ZipFile, sheet_path: str, ss):
    with z.open(sheet_path) as f:
        for event, elem in ET.iterparse(f, events=('end',)):
            if elem.tag==NS+'row':
                row={}
                for c in elem.findall(NS+'c'):
                    row[_col_idx(c.attrib.get('r',''))]=_cell_value(c, ss)
                yield row
                elem.clear()


def _num(v) -> float:
    if v is None or v=='' or v=='-': return 0.0
    s=str(v).replace(',','').replace('￥','').replace('¥','').strip()
    if s.endswith('%'):
        try: return float(s[:-1])/100
        except Exception: return 0.0
    try: return float(s)
    except Exception: return 0.0


def _norm(v) -> str:
    s='' if v is None else str(v).strip()
    return s if s and s!='-' else '未标注'


def _date_str(v):
    if v is None or v=='': return None
    s=str(v).strip().replace('/','-')
    try:
        x=float(s)
        if 30000<x<65000:
            return (datetime.datetime(1899,12,30)+datetime.timedelta(days=x)).date().isoformat()
    except Exception:
        pass
    m=re.match(r'(\d{4})-(\d{1,2})-(\d{1,2})', s)
    if m:
        y,mo,d=map(int,m.groups())
        try: return datetime.date(y,mo,d).isoformat()
        except Exception: return None
    return None


def _empty(): return defaultdict(float)

def _add(agg, key, vals):
    d=agg[key]
    for m in METRICS: d[m]+=vals.get(m,0.0)


def enrich(r: Dict[str,Any]):
    r['客单价']=round(r.get('支付金额',0)/r.get('支付买家数',0),2) if r.get('支付买家数',0) else 0
    r['支付转化率']=round(r.get('支付买家数',0)/r.get('商品访客数',0),4) if r.get('商品访客数',0) else 0
    r['加购率']=round(r.get('商品加购人数',0)/r.get('商品访客数',0),4) if r.get('商品访客数',0) else 0
    r['浏览深度']=round(r.get('商品浏览量',0)/r.get('商品访客数',0),2) if r.get('商品访客数',0) else 0
    r['退款率']=round(r.get('成功退款金额',0)/r.get('支付金额',0),4) if r.get('支付金额',0) else 0
    return r


def _final(agg, keys, sort_key=None, limit=None, desc=False):
    out=[]
    for key, vals in agg.items():
        if not isinstance(key, tuple): key=(key,)
        r={k:v for k,v in zip(keys,key)}
        for m in METRICS: r[m]=round(vals[m],2)
        enrich(r); out.append(r)
    if sort_key: out.sort(key=lambda x:x.get(sort_key,0), reverse=desc)
    if limit: out=out[:limit]
    return out


def _open_zip(src):
    if isinstance(src, (str, Path)):
        return zipfile.ZipFile(src)
    if isinstance(src, bytes):
        return zipfile.ZipFile(io.BytesIO(src))
    try:
        return zipfile.ZipFile(io.BytesIO(src.getvalue()))
    except Exception:
        return zipfile.ZipFile(io.BytesIO(src.read()))


def month_shift(month: str, delta: int) -> str:
    y,m=map(int, month.split('-'))
    m+=delta
    while m<1: y-=1; m+=12
    while m>12: y+=1; m-=12
    return f'{y:04d}-{m:02d}'


def parse_sales_workbook(src) -> Dict[str,Any]:
    by_day=defaultdict(_empty); by_month=defaultdict(_empty); by_channel=defaultdict(_empty); by_store=defaultdict(_empty); by_cat=defaultdict(_empty); by_model=defaultdict(_empty); by_style=defaultdict(_empty); by_product=defaultdict(_empty)
    total=defaultdict(float); sets={k:set() for k in ['channels','stores','categories','models']}
    min_date=max_date=None; rows_count=0; used_sheets=[]
    with _open_zip(src) as z:
        ss=_load_shared(z); paths=_sheet_paths(z)
        if not paths: raise ValueError('未找到数据源工作表。请确认 Excel 中包含「天猫数据源」或「京东抖音数据源」。')
        for sheet_name, sp in paths.items():
            rows=_iter_rows(z, sp, ss)
            try: header=next(rows)
            except StopIteration: continue
            headers={str(v).strip():k for k,v in header.items() if v is not None and str(v).strip()}
            if '统计日期' not in headers: continue
            used_sheets.append(sheet_name)
            def get(row,h):
                i=headers.get(h); return row.get(i) if i is not None else None
            for row in rows:
                ds=_date_str(get(row,'统计日期'))
                if not ds: continue
                rows_count+=1
                cat=_norm(get(row,'品类')); model=_norm(get(row,'型号')); channel=_norm(get(row,'渠道')); store=_norm(get(row,'店铺')); style=_norm(get(row,'款式'))
                product=_norm(get(row,'商品名称')); pid=_norm(get(row,'商品ID'))
                vals={m:_num(get(row,m)) for m in METRICS}
                if '成功退款金额' not in headers: vals['成功退款金额']=0.0
                for m,v in vals.items(): total[m]+=v
                sets['channels'].add(channel); sets['stores'].add(store); sets['categories'].add(cat); sets['models'].add(model)
                min_date=ds if min_date is None or ds<min_date else min_date
                max_date=ds if max_date is None or ds>max_date else max_date
                _add(by_day,(ds,channel,store,cat,model),vals); _add(by_month,(ds[:7],channel,store,cat,model),vals)
                _add(by_channel,channel,vals); _add(by_store,store,vals); _add(by_cat,cat,vals); _add(by_model,(model,channel,cat,store),vals); _add(by_style,(style,channel,cat,model),vals); _add(by_product,(product[:80],pid,channel,cat,model),vals)
    if rows_count == 0:
        raise ValueError('没有解析到有效数据行。请检查表头是否包含「统计日期」以及销售指标字段。\n如果您上传的是目标拆解 Excel（含"X年X月目标拆解及登记"工作表），请切换到左侧【销售目标】数据类型，不要使用【销售数据】上传。')
    tot={m:round(total[m],2) for m in METRICS}; enrich(tot)
    monthly=_final(by_month,['月份','渠道','店铺','品类','型号'],'月份')
    by_month_all=defaultdict(_empty)
    for r in monthly: _add(by_month_all,r['月份'],{m:r[m] for m in METRICS})
    all_months=_final(by_month_all,['月份'],'月份'); all_months.sort(key=lambda r:r['月份'])
    return {
        'meta':{'rows':rows_count,'dateRange':[min_date,max_date],'generatedAt':datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),'usedSheets':used_sheets},
        'filters':{k:sorted(v) for k,v in sets.items()}, 'totals':tot, 'daily':_final(by_day,['日期','渠道','店铺','品类','型号'],'日期'), 'monthly':monthly, 'all_months':all_months,
        'channels':_final(by_channel,['渠道'],'支付金额',None,True), 'stores':_final(by_store,['店铺'],'支付金额',None,True), 'categories':_final(by_cat,['品类'],'支付金额',None,True),
        'models':_final(by_model,['型号','渠道','品类','店铺'],'支付金额',600,True), 'styles':_final(by_style,['款式','渠道','品类','型号'],'支付金额',1000,True), 'products':_final(by_product,['商品名称','商品ID','渠道','品类','型号'],'支付金额',1200,True)
    }


def rows_to_csv(rows: List[Dict[str,Any]], cols: List[str]) -> bytes:
    out=io.StringIO(); w=csv.DictWriter(out, fieldnames=cols, extrasaction='ignore'); w.writeheader(); w.writerows(rows)
    return ('\ufeff'+out.getvalue()).encode('utf-8-sig')


def load_targets(file_bytes: bytes):
    """解析目标 Excel，返回 {月份: {'shop': [...], 'model': [...]}}"""
    import io, re
    import openpyxl as _xl
    from datetime import datetime as _dt, timedelta as _td

    wb = _xl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    targets = {}
    _excel_epoch = _dt(1899, 12, 30)

    for ws_name in wb.sheetnames:
        m = re.search(r'(\d+)年(\d+)月', ws_name)
        if not m:
            continue
        year = int('20' + m.group(1))
        month = int(m.group(2))
        ym = f'{year}-{month:02d}'
        ws = wb[ws_name]
        max_row = ws.max_row
        max_col = min(ws.max_column, 60)

        # ── 1. 找日期行 ──
        header_row = 0
        for r in range(1, min(20, max_row + 1)):
            c3 = ws.cell(r, 3).value
            c4 = ws.cell(r, 4).value
            if c3 == '店铺' and c4 == '指标':
                header_row = r
                break
        if header_row < 1:
            continue

        date_cols = []
        for candidate_row in (header_row - 1, header_row, header_row + 1):
            if candidate_row < 1:
                continue
            for c in range(7, max_col + 1):
                v = ws.cell(candidate_row, c).value
                dt = None
                if isinstance(v, (int, float)) and 40000 < v < 50000:
                    dt = _excel_epoch + _td(days=int(v))
                elif hasattr(v, 'strftime'):
                    dt = v
                if dt:
                    date_cols.append((c, dt.strftime('%Y-%m-%d')))
            if date_cols:
                break

        if not date_cols:
            continue

        shop_rows = []
        model_rows = []
        in_model_section = False
        current_shop = ''
        current_model = ''

        data_start = header_row + 1
        while data_start <= max_row:
            c3 = ws.cell(data_start, 3).value
            c4 = ws.cell(data_start, 4).value
            if c3 or c4:
                break
            data_start += 1
        for r in range(data_start, max_row + 1):
            c3 = ws.cell(r, 3).value
            c4 = ws.cell(r, 4).value
            c5 = ws.cell(r, 5).value

            if c3 and str(c3).strip() == '销售目标拆解':
                in_model_section = True
                continue
            if in_model_section and c3 and str(c3).strip() == '店铺' and c4 and str(c4).strip() == '型号':
                continue
            if not in_model_section:
                if not c3 and not c4:
                    continue
            else:
                if not c4 and not c5:
                    if c3 and str(c3).strip():
                        current_shop = str(c3).strip()
                    continue

            if not in_model_section:
                if c3 and str(c3).strip() == '合计':
                    current_shop = ''
                    continue
                if c3:
                    current_shop = str(c3).strip()
                if c4 and current_shop:
                    indicator = str(c4).strip()
                    row_data = {'店铺': current_shop, '指标': indicator}
                    for col_idx, date_str in date_cols:
                        v = ws.cell(r, col_idx).value
                        row_data[date_str] = float(v) if isinstance(v, (int, float)) else 0.0
                    e_val = ws.cell(r, 5).value
                    row_data['合计'] = float(e_val) if isinstance(e_val, (int, float)) else 0.0
                    shop_rows.append(row_data)
            else:
                if c3 and str(c3).strip():
                    shop_val = str(c3).strip()
                    if any(kw in shop_val for kw in ['推广', '小计', '合计', '总计']):
                        continue
                    current_shop = shop_val
                if c4 and str(c4).strip():
                    model_val = str(c4).strip()
                    if any(kw in model_val for kw in ['推广', '小计', '合计', '总计']):
                        continue
                    current_model = model_val
                if c5 and str(c5).strip():
                    indicator = str(c5).strip()
                    if indicator in ('小计', '合计', '总计'):
                        continue
                    if any(kw in indicator for kw in ['小计', '合计', '总计', '推广型号']):
                        continue
                    row_data = {'店铺': current_shop, '型号': current_model, '指标': indicator}
                    for col_idx, date_str in date_cols:
                        v = ws.cell(r, col_idx).value
                        row_data[date_str] = float(v) if isinstance(v, (int, float)) else 0.0
                    e_val = ws.cell(r, 5).value
                    excel_total = float(e_val) if isinstance(e_val, (int, float)) else 0.0
                    calc_total = sum(row_data[d] for d in [d for _, d in date_cols])
                    row_data['合计'] = calc_total if calc_total else excel_total
                    model_rows.append(row_data)

        targets[ym] = {
            'shop': shop_rows,
            'model': model_rows,
            'dates': [d for _, d in date_cols],
        }

    return targets


def load_promo_data(file_bytes: bytes):
    """Parse 京东推广数据源 + 天猫推广数据源 sheets"""
    import io
    wb = None
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        try:
            import xlrd
            wb = xlrd.open_workbook(file_name=None, file_contents=file_bytes)
        except Exception:
            return []
    rows = []
    for sheet_name in ['京东推广数据源', '天猫推广数据源']:
        try:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
        except Exception:
            continue
        if hasattr(ws, 'iter_rows'):
            all_rows = list(ws.iter_rows(values_only=True))
        else:
            all_rows = [ws.row_values(i) for i in range(ws.nrows)]
        if not all_rows:
            continue
        header = [str(c).strip() if c is not None else '' for c in all_rows[0]]
        for raw in all_rows[1:]:
            r = {}
            for i, h in enumerate(header):
                if i >= len(raw):
                    r[h] = ''
                else:
                    v = raw[i]
                    r[h] = v if v is not None else ''
            date_val = r.get('日期', '')
            if hasattr(date_val, 'strftime'):
                date_str = date_val.strftime('%Y-%m-%d')
            else:
                date_str = str(date_val)[:10]
            r['_date'] = date_str
            r['_店铺'] = r.get('店铺', '')
            r['_渠道'] = r.get('渠道', '')
            r['_品类'] = r.get('品类', '')
            r['_型号'] = r.get('型号', '')
            _scene = r.get('营销场景') or r.get('推广场景') or r.get('场景') or r.get('营销渠道') or ''
            r['_营销场景'] = str(_scene).strip() if _scene else r['_渠道']
            spend = r.get('花费', None) or r.get('花费', 0)
            r['_花费'] = float(spend) if spend not in (None, '') else 0.0
            impress = r.get('展现数', None) or r.get('展现数', 0)
            r['_展现数'] = float(impress) if impress not in (None, '') else 0.0
            clicks = r.get('点击数', None) or r.get('点击数', 0)
            r['_点击数'] = float(clicks) if clicks not in (None, '') else 0.0
            direct_amt = r.get('直接订单金额', None) or r.get('直接订单金额', 0)
            indirect_amt = r.get('间接订单金额', None) or r.get('间接订单金额', 0)
            total_amt = r.get('总订单金额', None) or r.get('总订单金额', 0)
            r['_直接订单金额'] = float(direct_amt) if direct_amt not in (None, '') else 0.0
            r['_间接订单金额'] = float(indirect_amt) if indirect_amt not in (None, '') else 0.0
            r['_总订单金额'] = float(total_amt) if total_amt not in (None, '') else 0.0
            r['_总加购数'] = float(r.get('总加购数', 0) or 0)
            cust = (r.get('成交客户数', None) or r.get('成交客户', None) or
                    r.get('订单客户数', None) or r.get('支付买家数', None) or
                    r.get('成交买家数', None) or 0)
            r['_成交客户数'] = float(cust) if cust not in (None, '') else 0.0
            total_orders = (r.get('总订单行', None) or r.get('订单行', None) or
                           r.get('成交订单数', None) or r.get('订单数', None) or
                           r.get('总成交订单数', None) or r.get('总订单数', None) or 0)
            r['_总成交订单量'] = float(total_orders) if total_orders not in (None, '') else 0.0
            direct_orders = (r.get('直接订单行', None) or r.get('直接成交订单数', None) or
                            r.get('直接订单数', None) or r.get('直接成交订单量', None) or 0)
            r['_直接订单量'] = float(direct_orders) if direct_orders not in (None, '') else 0.0
            roi = r.get('投产比', None) or r.get('投产比', 0)
            r['_投产比'] = float(roi) if roi not in (None, '') else 0.0
            rows.append(r)
    return rows
