import openpyxl as xl
from datetime import datetime as dt, timedelta as td

file_path = r'C:\Users\Gwell\WorkBuddy\2026-05-22-17-31-37\xiaotunbi\.data_cache\last_targets.xlsx'

wb = xl.load_workbook(file_path, data_only=True)

for ws_name in ['26年5月目标拆解及登记', '26年6月目标拆解及登记']:
    print(f'\n=== Sheet: {ws_name} ===')
    ws = wb[ws_name]
    max_row = ws.max_row
    max_col = min(ws.max_column, 60)

    header_row = 0
    for r in range(1, min(20, max_row + 1)):
        c3 = ws.cell(r, 3).value
        c4 = ws.cell(r, 4).value
        if c3 == '店铺' and c4 == '指标':
            header_row = r
            break

    date_cols = []
    for candidate_row in (header_row - 1, header_row, header_row + 1):
        if candidate_row < 1:
            continue
        for c in range(7, max_col + 1):
            v = ws.cell(candidate_row, c).value
            if v is not None:
                d = None
                if isinstance(v, (int, float)) and 40000 < v < 50000:
                    d = dt(1899, 12, 30) + td(days=int(v))
                elif hasattr(v, 'strftime'):
                    d = v
                if d:
                    date_cols.append((c, d.strftime('%Y-%m-%d')))
        if date_cols:
            break

    data_start = header_row + 1
    while data_start <= max_row:
        c3 = ws.cell(data_start, 3).value
        c4 = ws.cell(data_start, 4).value
        if c3 or c4:
            break
        data_start += 1

    # 单品区指标
    print('--- Model indicators ---')
    in_model = False
    model_indicators = set()
    for r in range(data_start, max_row + 1):
        c3 = ws.cell(r, 3).value
        c4 = ws.cell(r, 4).value
        c5 = ws.cell(r, 5).value
        if c3 and str(c3).strip() == '销售目标拆解':
            in_model = True
            continue
        if in_model and c5 and isinstance(c5, str) and c5 != '指标':
            model_indicators.add(c5)
    for ind in sorted(model_indicators):
        print(f'  {ind}')
